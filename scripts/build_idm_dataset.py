from __future__ import annotations

import json
import math
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BASES_DIR = ROOT / "bases"
SRC_DATA_DIR = ROOT / "src" / "data"
PUBLIC_DATA_DIR = ROOT / "public" / "data"
REGION_FILE = BASES_DIR / "RD_Mun.xlsx"


INDICATOR_METADATA = [
    {
        "key": "agua",
        "dimension": "ambiental",
        "label": "Indice de Atendimento Total de Agua",
        "sheet_label": "Índice de Atendimento Total de Água",
        "polarity": "positive",
        "referenceValue": "100%",
        "description": "% da populacao atendida com abastecimento de agua.",
        "source": "SNISA",
    },
    {
        "key": "esgoto",
        "dimension": "ambiental",
        "label": "Indice de Atendimento Total de Esgoto",
        "sheet_label": "Índice de Atendimento Total de Esgoto",
        "polarity": "positive",
        "referenceValue": "100%",
        "description": "% da populacao atendida com esgotamento sanitario em relacao a populacao atendida com agua.",
        "source": "SNISA",
    },
    {
        "key": "cobertura_vegetal",
        "dimension": "ambiental",
        "label": "Percentual de Cobertura Vegetal",
        "sheet_label": "Percentual de Cobertura Vegetal",
        "polarity": "positive",
        "referenceValue": "70%",
        "description": "Area com cobertura vegetal sobre a area total do municipio.",
        "source": "MapBiomas",
    },
    {
        "key": "focos_calor",
        "dimension": "ambiental",
        "label": "Densidade de Focos de Calor",
        "sheet_label": "Densidade de Focos de Calor",
        "polarity": "negative",
        "referenceValue": "1 foco/100 km²",
        "description": "Numero de focos de calor por 100 km².",
        "source": "INPE",
    },
    {
        "key": "gastos_gestao_ambiental",
        "dimension": "ambiental",
        "label": "Participacao dos Gastos do Municipio em Gestao Ambiental",
        "sheet_label": "Participação dos Gastos do Município em Gestão Ambiental",
        "polarity": "positive",
        "referenceValue": "3,2%",
        "description": "Percentual dos gastos em gestao ambiental sobre os gastos totais.",
        "source": "SICONFI",
    },
    {
        "key": "co2e_habitante",
        "dimension": "ambiental",
        "label": "Emissoes de CO2e por habitante",
        "sheet_label": "Emissões de CO2e por habitante",
        "polarity": "negative",
        "referenceValue": "0,72 t/hab",
        "description": "Emissoes brutas de dioxido de carbono equivalente por habitante.",
        "source": "SEEG e Condepe/Fidem",
    },
    {
        "key": "pib_per_capita",
        "dimension": "economia",
        "label": "PIB per Capita",
        "sheet_label": "PIB per Capita",
        "polarity": "positive",
        "referenceValue": "R$ 31.645,00",
        "description": "PIB municipal dividido pela populacao.",
        "source": "IBGE e Condepe/Fidem",
    },
    {
        "key": "formalizacao_trabalho",
        "dimension": "economia",
        "label": "Taxa de Formalizacao do Mercado de Trabalho",
        "sheet_label": "Taxa de Formalização do Mercado de Trabalho",
        "polarity": "positive",
        "referenceValue": "30%",
        "description": "Vinculos formais da RAIS sobre populacao em idade ativa.",
        "source": "RAIS e DATASUS",
    },
    {
        "key": "empregados_2sm",
        "dimension": "economia",
        "label": "Participacao de Empregados Formais com dois Salarios-Minimos ou mais",
        "sheet_label": "Participação de Empregados Formais com dois Salários-Mínimos ou mais",
        "polarity": "positive",
        "referenceValue": "35%",
        "description": "Percentual de vinculos formais recebendo 2 salarios-minimos ou mais.",
        "source": "RAIS",
    },
    {
        "key": "empregados_superior",
        "dimension": "economia",
        "label": "Participacao de Empregados Formais com Ensino Superior Completo",
        "sheet_label": "Participação de Empregados Formais com Ensino Superior Completo",
        "polarity": "positive",
        "referenceValue": "76%",
        "description": "Percentual de vinculos formais com ensino superior completo.",
        "source": "RAIS",
    },
    {
        "key": "servicos_vs_adm_publica",
        "dimension": "economia",
        "label": "Razao entre o Valor Adicionado dos Servicos e o Valor Adicionado da Administracao Publica",
        "sheet_label": "Razão entre o Valor Adicionado dos Serviços e o Valor Adicionado da Administração Pública",
        "polarity": "positive",
        "referenceValue": "2,1",
        "description": "Valor adicionado bruto dos servicos privados sobre o da administracao publica.",
        "source": "IBGE",
    },
    {
        "key": "banda_larga",
        "dimension": "economia",
        "label": "Densidade de Banda Larga",
        "sheet_label": "Densidade de Banda Larga",
        "polarity": "positive",
        "referenceValue": "10 por 100 hab",
        "description": "Acessos de banda larga por 100 habitantes.",
        "source": "Anatel",
    },
    {
        "key": "investimentos_publicos_pib",
        "dimension": "economia",
        "label": "Participacao do Valor de Investimentos Publicos no PIB",
        "sheet_label": "Participação do Valor de Investimentos Públicos no PIB",
        "polarity": "positive",
        "referenceValue": "3,6%",
        "description": "Percentual do investimento publico em relacao ao PIB.",
        "source": "SICONFI",
    },
    {
        "key": "pre_natal",
        "dimension": "social",
        "label": "Atendimento Pre-Natal",
        "sheet_label": "Atendimento Pré-Natal",
        "polarity": "positive",
        "referenceValue": "90%",
        "description": "Nascidos vivos com 7 ou mais consultas sobre o total de nascidos vivos.",
        "source": "DATASUS/SINASC",
    },
    {
        "key": "mortalidade_infantil",
        "dimension": "social",
        "label": "Taxa de Mortalidade Infantil",
        "sheet_label": "Taxa de Mortalidade Infantil",
        "polarity": "negative",
        "referenceValue": "8,5 por mil",
        "description": "Obitos de menores de 1 ano por mil nascidos vivos.",
        "source": "DATASUS/SINASC",
    },
    {
        "key": "mortalidade_avc",
        "dimension": "social",
        "label": "Taxa de Mortalidade por Acidente Vascular Cerebral (AVC)",
        "sheet_label": "Taxa de Mortalidade por Acidente Vascular Cerebral (AVC)",
        "polarity": "negative",
        "referenceValue": "27,1 obitos/100 mil hab",
        "description": "Obitos por AVC por 100 mil habitantes.",
        "source": "DATASUS/SIM",
    },
    {
        "key": "mortalidade_iam",
        "dimension": "social",
        "label": "Taxa de Mortalidade por Infarto Agudo do Miocardio (IAM)",
        "sheet_label": "Taxa de Mortalidade por Infarto Agudo do Miocárdio (IAM)",
        "polarity": "negative",
        "referenceValue": "30 obitos/100 mil hab",
        "description": "Obitos por IAM por 100 mil habitantes.",
        "source": "DATASUS/SIM",
    },
    {
        "key": "mortalidade_motos",
        "dimension": "social",
        "label": "Taxa de Mortalidade por Acidentes de Motos",
        "sheet_label": "Taxa de Mortalidade por Acidentes de Motos",
        "polarity": "negative",
        "referenceValue": "3,5 obitos/100 mil hab",
        "description": "Obitos por acidentes de moto por 100 mil habitantes.",
        "source": "DATASUS/SIM",
    },
    {
        "key": "mortalidade_agressao",
        "dimension": "social",
        "label": "Taxa de Mortalidade por Agressao",
        "sheet_label": "Taxa de Mortalidade por Agressão",
        "polarity": "negative",
        "referenceValue": "10 obitos/100 mil hab",
        "description": "Obitos por mortes violentas intencionais por 100 mil habitantes.",
        "source": "DATASUS/SIM",
    },
    {
        "key": "baixo_peso",
        "dimension": "social",
        "label": "Proporcao de Nascidos Vivos com Baixo Peso ao Nascer (<2.500 gramas)",
        "sheet_label": "Proporção de Nascidos Vivos com Baixo Peso ao Nascer (<2.500 gramas)",
        "polarity": "negative",
        "referenceValue": "4,6%",
        "description": "Percentual de nascidos vivos com baixo peso ao nascer.",
        "source": "DATASUS/SINASC",
    },
    {
        "key": "ideb_iniciais",
        "dimension": "social",
        "label": "Indice de Desenvolvimento da Educacao Basica - Anos Iniciais do Ensino Fundamental",
        "sheet_label": "Índice de Desenvolvimento da Educação Básica - Anos Iniciais do Ensino Fundamental",
        "polarity": "positive",
        "referenceValue": "7",
        "description": "Desempenho dos alunos do 1º ao 5º ano da rede municipal.",
        "source": "INEP/MEC",
    },
    {
        "key": "ideb_finais",
        "dimension": "social",
        "label": "Indice de Desenvolvimento da Educacao Basica - Anos Finais do Ensino Fundamental",
        "sheet_label": "Índice de Desenvolvimento da Educação Básica - Anos Finais do Ensino Fundamental",
        "polarity": "positive",
        "referenceValue": "6",
        "description": "Desempenho dos alunos do 6º ao 9º ano da rede municipal.",
        "source": "INEP/MEC",
    },
    {
        "key": "distorcao_iniciais",
        "dimension": "social",
        "label": "Taxa de Distorcao Idade-Serie nos Anos Iniciais do Ensino Fundamental",
        "sheet_label": "Taxa de Distorção Idade-Série nos Anos Iniciais do Ensino Fundamental",
        "polarity": "negative",
        "referenceValue": "4%",
        "description": "Percentual de alunos com atraso escolar maior que 2 anos nos anos iniciais.",
        "source": "INEP/MEC",
    },
    {
        "key": "distorcao_finais",
        "dimension": "social",
        "label": "Taxa de Distorcao Idade-Serie nos Anos Finais do Ensino Fundamental",
        "sheet_label": "Taxa de Distorção Idade-Série nos Anos Finais do Ensino Fundamental",
        "polarity": "negative",
        "referenceValue": "15%",
        "description": "Percentual de alunos com atraso escolar maior que 2 anos nos anos finais.",
        "source": "INEP/MEC",
    },
    {
        "key": "independencia_tributaria",
        "dimension": "governanca",
        "label": "Independencia Tributaria",
        "sheet_label": "Independência Tributária",
        "polarity": "positive",
        "referenceValue": "17%",
        "description": "Percentual da receita municipal proveniente de tributos locais.",
        "source": "SICONFI",
    },
    {
        "key": "complexidade_tributaria",
        "dimension": "governanca",
        "label": "Complexidade Tributaria",
        "sheet_label": "Complexidade Tributária",
        "polarity": "negative",
        "referenceValue": "0,35",
        "description": "Mede a diversificacao das fontes de receita tributaria.",
        "source": "SICONFI",
    },
    {
        "key": "capacidade_investimentos",
        "dimension": "governanca",
        "label": "Capacidade de Investimentos",
        "sheet_label": "Capacidade de Investimentos",
        "polarity": "positive",
        "referenceValue": "12%",
        "description": "Participacao da despesa de capital na despesa orcamentaria.",
        "source": "SICONFI",
    },
    {
        "key": "captacao_recursos",
        "dimension": "governanca",
        "label": "Captacao de Recursos",
        "sheet_label": "Captação de Recursos",
        "polarity": "positive",
        "referenceValue": "4,5%",
        "description": "Percentual de recursos captados em convenio sobre a receita corrente total.",
        "source": "SICONFI",
    },
]


DIMENSIONS = [
    {
        "key": "ambiental",
        "label": "Ambiental",
        "description": "Acesso a agua e esgoto, cobertura vegetal, focos de calor, gastos ambientais e emissoes.",
    },
    {
        "key": "economia",
        "label": "Economia",
        "description": "Renda, formalizacao, qualificacao, servicos privados, banda larga e investimento publico.",
    },
    {
        "key": "social",
        "label": "Social",
        "description": "Saude materno-infantil, mortalidade evitavel, seguranca e aprendizagem escolar.",
    },
    {
        "key": "governanca",
        "label": "Governanca Publica",
        "description": "Capacidade administrativa e financeira da gestao municipal.",
    },
]


CLASS_BANDS = [
    {"key": "classe_1", "label": "Classe 1", "min": 0.7, "max": None, "description": "Maior ou igual a 0,700"},
    {"key": "classe_2", "label": "Classe 2", "min": 0.6, "max": 0.7, "description": "Maior ou igual a 0,600 e menor que 0,700"},
    {"key": "classe_3", "label": "Classe 3", "min": 0.5, "max": 0.6, "description": "Maior ou igual a 0,500 e menor que 0,600"},
    {"key": "classe_4", "label": "Classe 4", "min": None, "max": 0.5, "description": "Menor que 0,500"},
]


def normalize_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    stripped = "".join(char for char in normalized if not unicodedata.combining(char))
    return stripped.lower().replace("'", "").replace("-", " ").replace("  ", " ").strip()


def to_number(value: Any) -> float | None:
    if value in (None, "", "Não Avaliado"):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_header_row(ws) -> tuple[int, int]:
    for row_index in range(1, 20):
        row = next(ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True))
        for col_index, value in enumerate(row):
            if value == "Município":
                return row_index, col_index
    raise RuntimeError(f"Header row not found for sheet {ws.title}")



def load_region_metadata() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(REGION_FILE, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    region_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        if not row or not row[3]:
            continue
        name = str(row[3]).strip()
        region_map[name] = {
            "ibgeCode": str(row[1]),
            "macro": str(row[7]).strip(),
            "macroCode": str(row[9]).strip(),
            "regional": str(row[10]).strip(),
            "latitude": float(row[12]) / 1000 if row[12] is not None else None,
            "longitude": float(row[13]) / 1000 if row[13] is not None else None,
        }
    return region_map

def load_consolidated() -> tuple[list[int], dict[str, dict[str, Any]]]:
    workbook = load_workbook(BASES_DIR / "Dados IDM-PE consolidados.xlsx", read_only=True, data_only=True)
    region_map = load_region_metadata()
    years = [2020, 2021, 2022, 2023]
    dimension_sheet_map = {
        "ambiental": "Ambiental",
        "economia": "Economia",
        "social": "Social",
        "governanca": "Governança Pública",
    }

    municipalities: dict[str, dict[str, Any]] = {}
    idm_sheet = workbook["IDM-PE"]

    summary_labels = {"M?dia", "Desv.padr?o", "Coef. de varia??o", "M?nimo", "M?ximo", "Mediana"}

    for row in idm_sheet.iter_rows(min_row=4, values_only=True):
        name = row[1]
        if not name or name in summary_labels:
            continue
        entry = municipalities.setdefault(
            name,
            {
                "name": name,
                "slug": normalize_name(name),
                "region": region_map.get(name, {}),
                "idm": {},
                "dimensions": {dimension["key"]: {} for dimension in DIMENSIONS},
                "indicators": {},
            },
        )
        for offset, year in enumerate(years, start=2):
            entry["idm"][str(year)] = to_number(row[offset])

    for dimension_key, sheet_name in dimension_sheet_map.items():
        ws = workbook[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = row[1]
            if not name or name in summary_labels:
                continue
            entry = municipalities.get(name)
            if not entry:
                continue
            for offset, year in enumerate(years, start=2):
                entry["dimensions"][dimension_key][str(year)] = to_number(row[offset])

    return years, municipalities


def load_standardized(municipalities: dict[str, dict[str, Any]]) -> None:
    workbook = load_workbook(BASES_DIR / "Dados IDM-PE padronizados.xlsx", read_only=True, data_only=True)
    label_to_metadata = {item["sheet_label"]: item for item in INDICATOR_METADATA}

    for year_name in workbook.sheetnames:
        ws = workbook[year_name]
        header_row_index, municipality_column = find_header_row(ws)
        header_row = next(ws.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))
        indicator_columns = []

        for col_index in range(municipality_column + 1, len(header_row)):
            header = header_row[col_index]
            if header is None:
                continue
            metadata = label_to_metadata.get(header)
            if metadata:
                indicator_columns.append((col_index, metadata))

        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            municipality_name = row[municipality_column]
            if not municipality_name:
                continue
            entry = municipalities.get(municipality_name)
            if not entry:
                continue
            entry["indicators"].setdefault(year_name, {})
            for col_index, metadata in indicator_columns:
                entry["indicators"][year_name][metadata["key"]] = to_number(row[col_index])


def class_for_value(value: float | None) -> str | None:
    if value is None:
        return None
    if value >= 0.7:
        return "classe_1"
    if value >= 0.6:
        return "classe_2"
    if value >= 0.5:
        return "classe_3"
    return "classe_4"



def build_region_summary(years: list[int], municipalities: dict[str, dict[str, Any]]) -> dict[str, Any]:
    region_summary: dict[str, Any] = {"macro": {}, "regional": {}}
    for level in ["macro", "regional"]:
        grouped: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
        for municipality in municipalities.values():
            region_name = municipality.get("region", {}).get(level)
            if not region_name:
                continue
            for year in years:
                value = municipality["idm"].get(str(year))
                if value is not None:
                    grouped[region_name][str(year)].append(value)
        region_summary[level] = {
            region_name: {year: sum(values) / len(values) for year, values in years_map.items()}
            for region_name, years_map in grouped.items()
        }
    return region_summary

def build_summary(years: list[int], municipalities: dict[str, dict[str, Any]]) -> dict[str, Any]:
    summary: dict[str, Any] = {"years": {}}
    for year in years:
        year_key = str(year)
        values = []
        counts = defaultdict(int)
        for municipality in municipalities.values():
            value = municipality["idm"].get(year_key)
            if value is None:
                continue
            values.append((municipality["name"], value))
            band = class_for_value(value)
            if band:
                counts[band] += 1
        values.sort(key=lambda item: item[1], reverse=True)
        mean = sum(item[1] for item in values) / len(values) if values else None
        summary["years"][year_key] = {
            "top": {"name": values[0][0], "value": values[0][1]} if values else None,
            "bottom": {"name": values[-1][0], "value": values[-1][1]} if values else None,
            "average": mean,
            "classCounts": counts,
        }
    return summary


def write_geojson_copy() -> None:
    source = BASES_DIR / "geojs-26-mun.json"
    content = source.read_text(encoding="utf-8")
    (PUBLIC_DATA_DIR / "pe-municipios.geojson").write_text(content, encoding="utf-8")
    (SRC_DATA_DIR / "peMunicipios.json").write_text(content, encoding="utf-8")


def main() -> None:
    SRC_DATA_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_DATA_DIR.mkdir(parents=True, exist_ok=True)

    years, municipalities = load_consolidated()
    load_standardized(municipalities)
    summary = build_summary(years, municipalities)
    region_summary = build_region_summary(years, municipalities)

    dataset = {
        "title": "Indice de Desenvolvimento Municipal de Pernambuco",
        "subtitle": "Painel exploratorio montado a partir das bases consolidadas e padronizadas do IDM-PE.",
        "years": years,
        "dimensions": DIMENSIONS,
        "indicatorMetadata": INDICATOR_METADATA,
        "classBands": CLASS_BANDS,
        "methodology": {
            "period": "2020 a 2023",
            "note": "O IDM-PE agrega 28 indicadores em quatro dimensoes. Os indicadores sao normalizados entre 0 e 1 com valores de referencia fixos para garantir comparabilidade intertemporal. As dimensoes usam media aritmetica e o IDM final usa media geometrica.",
        },
        "summary": summary,
        "regionSummary": region_summary,
        "municipalities": sorted(municipalities.values(), key=lambda item: item["name"]),
    }

    serialized = json.dumps(dataset, ensure_ascii=False, separators=(",", ":"))
    (SRC_DATA_DIR / "idmDataset.json").write_text(serialized, encoding="utf-8")
    (PUBLIC_DATA_DIR / "idmDataset.json").write_text(serialized, encoding="utf-8")
    write_geojson_copy()


if __name__ == "__main__":
    main()
